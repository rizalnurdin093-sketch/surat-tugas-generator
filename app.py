"""
app.py — Entry point aplikasi Surat Tugas Generator

Route yang ada:
  GET  /            -> halaman utama (isi judul/tempat/tanggal + upload Excel)
  GET/POST /login   -> login
  GET  /logout      -> logout
  POST /preview     -> upload Excel, parse, tampilkan preview
  GET  /riwayat     -> daftar surat yang pernah dibuat

Struktur:
  - app.py           : semua route + logika web
  - config.py        : konfigurasi
  - database.py      : semua query database
  - parser_excel.py  : logika baca Excel
  - templates/       : halaman HTML (base.html = header/footer sekali)
  - static/          : css & js
"""

import os
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user

from config import Config
import database
from parser_excel import parse_excel, sort_az

# ── Setup Flask ─────────────────────────────────────────────────────
app = Flask(__name__)
app.config.from_object(Config)

# Pastikan folder yang dibutuhkan ada
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(app.config["GENERATED_FOLDER"], exist_ok=True)
os.makedirs(app.config["TEMPLATE_SURAT_FOLDER"], exist_ok=True)

# ── Setup login ─────────────────────────────────────────────────────
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"


class User:
    """Wrapper user untuk flask-login (baca dari sqlite)."""

    def __init__(self, id, username):
        self.id = id
        self.username = username
        self.is_authenticated = True
        self.is_active = True
        self.is_anonymous = False

    def get_id(self):
        return str(self.id)


@login_manager.user_loader
def load_user(user_id):
    # Flask-login butuh is_authenticated dll, tapi kita pakai User class sederhana.
    # Ambil dari database berdasarkan id.
    # Catatan: kita simpan username di session, id juga valid.
    row = _find_user_by_id(int(user_id))
    if row:
        return User(row["id"], row["username"])
    return None


def _find_user_by_id(user_id):
    import sqlite3
    conn = database.get_db()
    row = conn.execute(
        "SELECT * FROM users WHERE id = ?", (user_id,)
    ).fetchone()
    conn.close()
    return row


# ── Helper ──────────────────────────────────────────────────────────
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in app.config["ALLOWED_EXTENSIONS"]


# ── Routes ──────────────────────────────────────────────────────────
@app.route("/")
@login_required
def index():
    return render_template("index.html", username=current_user.username)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        user = database.get_user(username)
        if user and database.verify_password(user, password):
            login_user(User(user["id"], user["username"]))
            return redirect(url_for("index"))
        else:
            flash("Username atau password salah.", "error")
    return render_template("login.html")


@app.route("/logout", methods=["GET", "POST"])
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


@app.route("/preview", methods=["POST"])
@login_required
def preview():
    if "file" not in request.files:
        flash("Pilih file Excel dulu.", "error")
        return redirect(url_for("index"))

    file = request.files["file"]
    if file.filename == "":
        flash("Nama file kosong.", "error")
        return redirect(url_for("index"))

    if not allowed_file(file.filename):
        flash("Format file harus .xlsx (tidak support .xls lama). Simpan ulang sebagai .xlsx.", "error")
        return redirect(url_for("index"))

    # Simpan file upload
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
    file.save(filepath)

    try:
        peserta = parse_excel(filepath)
        if not peserta:
            flash("File tidak berisi data peserta yang valid.", "error")
            return redirect(url_for("index"))
    except ValueError as e:
        flash(str(e), "error")
        return redirect(url_for("index"))

    peserta = sort_az(peserta)

    # Simpan data sementara di session: cukup nama file (kecil).
    # Data peserta tidak disimpan di cookie — bisa besar & melewati limit 4KB.
    # File Excel sudah disimpan di uploads/, nanti dibaca ulang di generate.
    session["uploaded_filename"] = file.filename
    session["form"] = {
        "judul": request.form.get("judul", "").strip(),
        "tempat": request.form.get("tempat", "").strip(),
        "tanggal": request.form.get("tanggal", "").strip(),
        "tanggal_surat": request.form.get("tanggal_surat", "").strip(),
    }

    return render_template(
        "index.html",
        username=current_user.username,
        peserta=peserta,
        total=len(peserta),
        file_name=file.filename,
        form_data=session["form"],
    )


@app.route("/generate", methods=["POST"])
@login_required
def generate():
    """Generate surat .docx dari data yang tersimpan di session (filter+form)."""
    from generator_surat import generate_surat

    filename = session.get("uploaded_filename")
    form = session.get("form", {})

    if not filename:
        flash("Upload file Excel dulu.", "error")
        return redirect(url_for("index"))

    # Baca ulang peserta dari file yang sudah diupload (tidak disimpan di cookie session)
    from parser_excel import parse_excel, sort_az
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    if not os.path.exists(filepath):
        flash("File tidak ditemukan. Upload ulang.", "error")
        return redirect(url_for("index"))
    peserta = sort_az(parse_excel(filepath))

    judul = form.get("judul", "Pelatihan")
    tempat = form.get("tempat", "")
    tanggal = form.get("tanggal", "")
    tanggal_surat = form.get("tanggal_surat", "")

    if not judul or not tempat or not tanggal:
        flash("Isi dulu Judul, Tempat, dan Tanggal pelatihan.", "error")
        return redirect(url_for("index"))

    try:
        output_path, filename = generate_surat(
            judul=judul,
            tempat=tempat,
            hari_tanggal=tanggal,
            tanggal_surat=tanggal_surat,
            peserta=peserta,
            output_dir=app.config["GENERATED_FOLDER"],
        )

        # Simpan riwayat
        database.add_surat_history(
            judul=judul,
            tempat=tempat,
            tanggal=tanggal,
            jumlah_peserta=len(peserta),
            file_name=filename,
            file_peserta=session.get("uploaded_filename"),
            tanggal_surat=tanggal_surat,
            dibuat_oleh=current_user.username,
        )

        return send_file(output_path, as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f"Gagal generate surat: {str(e)}", "error")
        return redirect(url_for("index"))


# ── Master Data Surat ────────────────────────────────────────────────
@app.route("/master")
@login_required
def master():
    """Halaman master data surat: daftar + search + export + atur nomor + download."""
    from urllib.parse import quote
    q = request.args.get("q", "").strip()
    if q:
        history = database.search_surat_history(q)
    else:
        history = database.get_all_surat_history()
    return render_template(
        "master.html",
        username=current_user.username,
        history=history,
        search_q=q,
    )


@app.route("/master/atur/<int:surat_id>", methods=["POST"])
@login_required
def master_atur_nomor(surat_id):
    """Update nomor surat (angka saja, misal '2630')."""
    nomor = request.form.get("nomor_angka", "").strip()
    if not nomor.isdigit():
        flash("Nomor harus angka (contoh: 2630).", "error")
        return redirect(url_for("master"))
    database.update_nomor_surat(surat_id, nomor)
    flash(f"Nomor surat diupdate: {nomor}/Tbk/ST-4010/26-S8.7.1", "success")
    return redirect(url_for("master"))


def _parse_tanggal_mulai_selesai(tanggal_str):
    """
    Parse string tanggal pelatihan menjadi (mulai, selesai).

    Format contoh: 'Senin-Selasa / 13 Agustus 2026 - 14 Agustus 2026'
      -> ('13 Agustus 2026', '14 Agustus 2026')
    Jika hanya satu tanggal: 'Senin / 13 Agustus 2026'
      -> ('13 Agustus 2026', '13 Agustus 2026')
    Jika tak bisa diparse: kembalikan (string_asli, '').
    """
    import re
    s = (tanggal_str or "").strip()
    # Cari pola dua tanggal: "13 Agustus 2026 - 14 Agustus 2026"
    # Tanggal = angka + kata bulan + 4 digit tahun. Tanggal mulai = angka saja (1-2 digit).
    pat = r"(\d{1,2}\s+[A-Za-z]+\.?\s+\d{4})\s*[-–—]\s*(\d{1,2}\s+[A-Za-z]+\.?\s+\d{4})"
    m = re.search(pat, s)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    # Coba satu tanggal saja
    pat1 = r"(\d{1,2}\s+[A-Za-z]+\.?\s+\d{4})"
    m1 = re.search(pat1, s)
    if m1:
        t = m1.group(1).strip()
        return t, t
    return s, ""


@app.route("/master/export")
@login_required
def master_export():
    """Download semua master data sebagai .xlsx (laporan)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    import io

    history = database.get_all_surat_history()

    wb = Workbook()
    ws = wb.active
    ws.title = "Master Surat Tugas"

    # Header (urutan tetap: ID, Judul, Nomor, Tempat, Tanggal Mulai, Tanggal Selesai)
    headers = [
        "ID", "Judul Pelatihan", "Nomor Surat", "Tempat",
        "Tanggal Mulai", "Tanggal Selesai"
    ]
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Data
    for i, r in enumerate(history, 2):
        nomor = r["nomor_surat"]
        if nomor and not nomor.startswith("/Tbk"):
            nomor = f"{nomor}/Tbk/ST-4010/26-S8.7.1"
        elif not nomor:
            nomor = "/Tbk/ST-4010/26-S8.7.1"

        tanggal_mulai, tanggal_selesai = _parse_tanggal_mulai_selesai(r["tanggal"])

        row_data = [
            r["id"],
            r["judul"],
            nomor,
            r["tempat"],
            tanggal_mulai,
            tanggal_selesai,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Auto-fit columns (semi)
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name=f"master_surat_tugas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Jalankan ────────────────────────────────────────────────────────
with app.app_context():
    database.init_db()


def create_dispatched_app():
    """
    Bungkus Flask app supaya jalan di subpath /surat/ (di belakang Nginx reverse proxy).
    Semua url_for otomatis dapat prefix /surat, jadi login/static/CSS tetap benar.
    """
    from werkzeug.middleware.dispatcher import DispatcherMiddleware
    from werkzeug.serving import run_simple

    application = DispatcherMiddleware(Flask(__name__).wsgi_app, {"/surat": app.wsgi_app})
    run_simple("0.0.0.0", 5001, application, use_reloader=False)


if __name__ == "__main__":
    # Dev biasa di root (tanpa subpath) — cepat untuk ngetes langsung di localhost
    app.run(host="0.0.0.0", port=5001, debug=True)

