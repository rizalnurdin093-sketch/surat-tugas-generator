"""
parser_excel.py — Logika membaca file Excel peserta.

File Excel kantor punya banyak kolom & beberapa sheet. Yang kita butuh cuma:
  - Nama   (kolom Employee Name / NAMA)
  - NIK    (kolom Employee No. / NIK)
  - Divisi (kolom Division / DIVISION / DIVISI)

Parser ini cari kolom secara fleksibel (cocokkan nama, abaikan huruf besar/kecil),
dan cuma ambil baris yang layak (nama + nik tidak kosong).

Dipisah dari app.py supaya gampang dites & dilacak kalau ada bug.
"""

import openpyxl


def _clean(text):
    """Bersihkan teks: ganti non-breaking space (\\xa0) dengan spasi biasa, rapikan."""
    if text is None:
        return ""
    return " ".join(str(text).replace("\xa0", " ").split())


def _normalize(text):
    """Bersihkan teks header untuk pencocokan (hapus spasi/simbol, lowercase)."""
    return "".join(ch.lower() for ch in _clean(text) if ch.isalnum())


# Peta nama kolom target -> daftar kemungkinan nama kolom di file
COLUMN_ALIASES = {
    "nama": ["employeename", "nama", "name"],
    "nik": ["employeeno", "nik", "employeenumber", "noid", "nomerinduk"],
    "divisi": ["division", "divisi", "department", "organisasi"],
}


def _find_column_indices(header_row):
    """
    Temukan index kolom Nama/NIK/Divisi di baris header.
    Header bisa di baris mana pun — kita cari baris yang mengandung
    kolom 'nama' + 'nik' bersamaan.
    """
    best = None
    for row_idx, row in enumerate(header_row):
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            key = _normalize(cell)
            # Cek apakah cell ini cocok untuk salah satu target
            for target, aliases in COLUMN_ALIASES.items():
                if key in aliases:
                    if best is None:
                        best = {"header_row": row_idx, "cols": {}, "score": 0}
                    # jangan overwrite kalau sudah dapat kolom itu
                    if target not in best["cols"]:
                        best["cols"][target] = col_idx
                        best["score"] += 1

        # Kalau di satu baris sudah ketemu nama+nik+divisi, pakai baris ini
        if best is not None:
            cols = best["cols"]
            if "nama" in cols and "nik" in cols and "divisi" in cols:
                return best["header_row"], cols

    # Kalau tidak nemu semua, return yang terbaik yang ada (setidaknya nama+nik)
    if best is not None and "nama" in best["cols"] and "nik" in best["cols"]:
        return best["header_row"], best["cols"]
    return None, {}


def parse_excel(file_path):
    """
    Baca file Excel, ambil data Nama/NIK/Divisi.
    Return list of dict: [{'nama':..., 'nik':..., 'divisi':...}, ...]
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Pilih sheet: prefer 'Peserta Fix', kalau tidak ada pakai sheet pertama
    sheet_name = "Peserta Fix" if "Peserta Fix" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    header_row_idx, cols = _find_column_indices(rows)
    if header_row_idx is None or not cols:
        raise ValueError(
            "Tidak bisa menemukan kolom Nama/NIK/Divisi di file. "
            "Pastikan file memiliki kolom (Employee Name / NAMA), "
            "(Employee No. / NIK), dan (Division / DIVISION)."
        )

    nama_col = cols["nama"]
    nik_col = cols["nik"]
    divisi_col = cols.get("divisi")

    peserta = []
    # Mulai dari baris setelah header (idx 0 = header)
    for row in rows[header_row_idx + 1:]:
        if row is None:
            continue
        nama = row[nama_col] if nama_col < len(row) else None
        nik = row[nik_col] if nik_col < len(row) else None
        divisi = row[divisi_col] if divisi_col is not None and divisi_col < len(row) else None

        # Lewati baris kosong / tanpa nama & nik
        if not nama or not nik:
            continue

        peserta.append({
            "nama": _clean(nama),
            "nik": _clean(nik),
            "divisi": _clean(divisi),
        })

    return peserta


def sort_az(peserta):
    """Urutkan daftar peserta berdasarkan nama A-Z (abaikan huruf besar/kecil)."""
    return sorted(peserta, key=lambda p: p["nama"].lower())
